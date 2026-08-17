"""
笨笨鱼财务系统 - Excel 导出工具
使用 openpyxl 将账本数据导出为 Excel 文件
"""

from io import BytesIO
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def _style_header(ws, headers: list[str]):
    """设置表头样式"""
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def _auto_column_width(ws):
    """自动调整列宽"""
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 4, 40)


def export_personal_records(records: list, stats: dict) -> BytesIO:
    """
    导出个人用户账本为 Excel
    :param records: 记录列表（ORM 对象或字典）
    :param stats: 统计数据字典
    :return: Excel 文件字节流
    """
    wb = Workbook()

    # ---- 工作表1：收支明细 ----
    ws_detail = wb.active
    ws_detail.title = "收支明细"

    headers = ["序号", "日期", "类型", "分类", "金额", "支付方式", "备注"]
    _style_header(ws_detail, headers)

    type_map = {"income": "收入", "expense": "支出"}

    for idx, record in enumerate(records, 1):
        ws_detail.append([
            idx,
            str(record.record_date),
            type_map.get(record.record_type, record.record_type),
            record.category,
            float(record.amount),
            record.payment_method,
            record.description or "",
        ])

    _auto_column_width(ws_detail)

    # ---- 工作表2：统计汇总 ----
    ws_stats = wb.create_sheet("统计汇总")
    ws_stats.append(["项目", "金额"])
    ws_stats.append(["总收入", float(stats.get("total_income", 0))])
    ws_stats.append(["总支出", float(stats.get("total_expense", 0))])
    ws_stats.append(["结余", float(stats.get("balance", 0))])
    ws_stats.append(["记录条数", stats.get("record_count", 0)])

    if stats.get("category_summary"):
        ws_stats.append([])
        ws_stats.append(["分类", "类型", "合计金额", "笔数"])
        for item in stats["category_summary"]:
            ws_stats.append([
                item["category"],
                type_map.get(item["record_type"], item["record_type"]),
                float(item["total"]),
                item["count"],
            ])

    _auto_column_width(ws_stats)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_org_records(records: list, stats: dict) -> BytesIO:
    """
    导出单位用户财政账本为 Excel
    :param records: 记录列表
    :param stats: 统计数据字典
    :return: Excel 文件字节流
    """
    wb = Workbook()

    # ---- 工作表1：业务明细 ----
    ws_detail = wb.active
    ws_detail.title = "财政业务明细"

    headers = [
        "序号", "日期", "类型", "凭证号", "预算科目编码", "部门",
        "项目名称", "资金来源", "经济分类", "功能分类",
        "金额", "收款方/付款方", "经办人", "审批人", "摘要",
    ]
    _style_header(ws_detail, headers)

    type_map = {"income": "收入", "expense": "支出"}

    for idx, record in enumerate(records, 1):
        ws_detail.append([
            idx,
            str(record.record_date),
            type_map.get(record.record_type, record.record_type),
            record.voucher_no or "",
            record.budget_code or "",
            record.department or "",
            record.project_name or "",
            record.fund_source or "",
            record.economic_classification or "",
            record.functional_classification or "",
            float(record.amount),
            record.payee_payer or "",
            record.handler or "",
            record.approver or "",
            record.description or "",
        ])

    _auto_column_width(ws_detail)

    # ---- 工作表2：统计汇总 ----
    ws_stats = wb.create_sheet("统计汇总")
    ws_stats.append(["项目", "金额"])
    ws_stats.append(["总收入", float(stats.get("total_income", 0))])
    ws_stats.append(["总支出", float(stats.get("total_expense", 0))])
    ws_stats.append(["结余", float(stats.get("balance", 0))])
    ws_stats.append((["记录条数", stats.get("record_count", 0)]))

    if stats.get("department_summary"):
        ws_stats.append([])
        ws_stats.append(["部门", "收入", "支出", "笔数"])
        for item in stats["department_summary"]:
            ws_stats.append([
                item["department"],
                float(item.get("income", 0)),
                float(item.get("expense", 0)),
                item["count"],
            ])

    if stats.get("fund_source_summary"):
        ws_stats.append([])
        ws_stats.append(["资金来源", "合计金额", "笔数"])
        for item in stats["fund_source_summary"]:
            ws_stats.append([
                item["fund_source"],
                float(item["total"]),
                item["count"],
            ])

    _auto_column_width(ws_stats)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
